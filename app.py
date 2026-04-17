import eventlet
eventlet.monkey_patch()
import json
import os
import re
import time
import hashlib
import uuid
from datetime import datetime
from functools import wraps
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
from flask_socketio import SocketIO, emit
import io

# ---------------------------------------------------------------------------
# Admin Credentials (change before production!)
# ---------------------------------------------------------------------------
ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin1234')

# ---------------------------------------------------------------------------
# App & SocketIO Setup
# ---------------------------------------------------------------------------
app = Flask(__name__)
app.secret_key = 'visionverse-secret-key-2024'
app.config['SESSION_TYPE'] = 'filesystem'
socketio = SocketIO(app, cors_allowed_origins="*")

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')

# ---------------------------------------------------------------------------
# Gesture engine (optional – works without camera hardware)
# ---------------------------------------------------------------------------
try:
    from gesture_engine import GestureEngine
    gesture_engine = GestureEngine(socketio)
    GESTURE_AVAILABLE = True
except Exception:
    gesture_engine = None
    GESTURE_AVAILABLE = False

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_json(filename):
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        return {} if filename == 'students.json' else []
    with open(path, 'r') as f:
        return json.load(f)

def _write_json(filename, data):
    path = os.path.join(DATA_DIR, filename)
    with open(path, 'w') as f:
        json.dump(data, f, indent=2)

def _hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'student_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('is_admin'):
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated

# ---------------------------------------------------------------------------
# AUTH ROUTES
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    if 'student_id' in session:
        return redirect(url_for('home'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'GET':
        return render_template('register.html')

    data = request.json
    students = _read_json('students.json')

    # Validate
    required = ['name', 'email', 'password', 'confirm_password', 'disability_type', 'dob', 'phone']
    for field in required:
        if not data.get(field, '').strip():
            return jsonify({'status': 'error', 'message': f'{field} is required'}), 400

    if data['password'] != data['confirm_password']:
        return jsonify({'status': 'error', 'message': 'Passwords do not match'}), 400

    # Auto-generate student_id
    max_id = 0
    for key in students.keys():
        if key.startswith('STU'):
            try:
                num = int(key[3:])
                if num > max_id:
                    max_id = num
            except ValueError:
                pass
    sid = f"STU{max_id + 1:03d}"

    # Check email uniqueness
    for s in students.values():
        if s.get('email', '').lower() == data['email'].strip().lower():
            return jsonify({'status': 'error', 'message': 'Email already registered'}), 400

    students[sid] = {
        'id': sid,
        'name': data['name'].strip(),
        'email': data['email'].strip().lower(),
        'password_hash': _hash_password(data['password']),
        'disability_type': data['disability_type'].strip(),
        'dob': data['dob'].strip(),
        'phone': data['phone'].strip(),
        'registered_at': datetime.now().isoformat()
    }
    _write_json('students.json', students)
    return jsonify({'status': 'success', 'message': f'Registration successful. Your Student ID is {sid}', 'student_id': sid})

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        if 'student_id' in session:
            return redirect(url_for('home'))
        return render_template('login.html')

    data = request.json
    identifier = data.get('identifier', '').strip()
    password = data.get('password', '')
    students = _read_json('students.json')

    # Find by student_id or email
    student = None
    if identifier in students:
        student = students[identifier]
    else:
        for s in students.values():
            if s.get('email', '').lower() == identifier.lower():
                student = s
                break

    if not student or student['password_hash'] != _hash_password(password):
        return jsonify({'status': 'error', 'message': 'Invalid credentials'}), 401

    session['student_id'] = student['id']
    session['student_name'] = student['name']
    session['student_email'] = student['email']
    session['disability_type'] = student['disability_type']
    session['exam_mode'] = None
    session['exam_status'] = 'Not Started'
    return jsonify({'status': 'success', 'message': 'Login successful'})

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ---------------------------------------------------------------------------
# PAGE ROUTES (login required)
# ---------------------------------------------------------------------------

@app.route('/home')
@login_required
def home():
    return render_template('home.html')

@app.route('/permissions')
@login_required
def permissions():
    return render_template('permissions.html')

@app.route('/exam')
@login_required
def exam():
    session['exam_status'] = 'In Progress'
    return render_template('exam.html')

@app.route('/results')
@login_required
def results():
    return render_template('results.html')

# ---------------------------------------------------------------------------
# API ROUTES
# ---------------------------------------------------------------------------

@app.route('/api/questions')
@login_required
def api_questions():
    questions = _read_json('questions.json')
    # Strip correct answers for client
    safe = []
    for q in questions:
        safe.append({
            'id': q['id'],
            'question': q['question'],
            'options': q['options']
        })
    return jsonify(safe)

@app.route('/api/submit', methods=['POST'])
@login_required
def api_submit():
    data = request.json
    answers = data.get('answers', {})  # {question_id: "A"/"B"/"C"/"D"}
    time_taken = data.get('time_taken', 0)

    questions = _read_json('questions.json')
    results_data = _read_json('results.json')

    score = 0
    breakdown = []
    for q in questions:
        qid = str(q['id'])
        student_answer = answers.get(qid, '')
        is_correct = student_answer == q['correct']
        if is_correct:
            score += 1
        breakdown.append({
            'id': q['id'],
            'question': q['question'],
            'options': q['options'],
            'correct': q['correct'],
            'student_answer': student_answer,
            'is_correct': is_correct
        })

    total = len(questions)
    pct = round((score / total) * 100) if total > 0 else 0

    if pct >= 80:
        badge = 'Excellent'
    elif pct >= 50:
        badge = 'Good'
    else:
        badge = 'Needs Improvement'

    result = {
        'result_id': str(uuid.uuid4())[:8],
        'student_id': session['student_id'],
        'student_name': session['student_name'],
        'exam_mode': session.get('exam_mode', 'unknown'),
        'score': score,
        'total': total,
        'percentage': pct,
        'badge': badge,
        'time_taken': time_taken,
        'breakdown': breakdown,
        'submitted_at': datetime.now().isoformat()
    }

    results_data.append(result)
    _write_json('results.json', results_data)

    session['exam_status'] = 'Completed'
    session['last_result'] = result

    return jsonify({'status': 'success', 'result': result})

@app.route('/api/student')
@login_required
def api_student():
    """Return current session student info for sidebar."""
    return jsonify({
        'student_id': session.get('student_id'),
        'name': session.get('student_name'),
        'email': session.get('student_email'),
        'disability_type': session.get('disability_type'),
        'exam_mode': session.get('exam_mode'),
        'exam_status': session.get('exam_status', 'Not Started')
    })

@app.route('/api/set_mode', methods=['POST'])
@login_required
def api_set_mode():
    mode = request.json.get('mode')
    session['exam_mode'] = mode
    return jsonify({'status': 'success'})

# ---------------------------------------------------------------------------
# VIDEO FEED
# ---------------------------------------------------------------------------

@app.route('/video_feed')
def video_feed():
    if gesture_engine and GESTURE_AVAILABLE:
        gesture_engine.start()  # auto-start so we don't depend on socket timing
        return Response(gesture_engine.generate_frames(),
                        mimetype='multipart/x-mixed-replace; boundary=frame')
    return Response(status=204)

# ---------------------------------------------------------------------------
# SOCKETIO EVENTS
# ---------------------------------------------------------------------------

@socketio.on('toggle_camera')
def handle_camera(data):
    if gesture_engine:
        if data.get('action') == 'start':
            gesture_engine.start()
        else:
            gesture_engine.stop()

@socketio.on('connect')
def handle_connect():
    pass

# ---------------------------------------------------------------------------
# ADMIN ROUTES
# ---------------------------------------------------------------------------

@app.route('/admin')
def admin_index():
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard'))
    return redirect(url_for('admin_login'))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if session.get('is_admin'):
        return redirect(url_for('admin_dashboard'))
    if request.method == 'GET':
        return render_template('admin_login.html')
    data = request.json
    if (data.get('username', '').strip() == ADMIN_USERNAME and
            data.get('password', '') == ADMIN_PASSWORD):
        session['is_admin'] = True
        return jsonify({'status': 'success'})
    return jsonify({'status': 'error', 'message': 'Invalid admin credentials'}), 401

@app.route('/admin/logout')
def admin_logout():
    session.pop('is_admin', None)
    return redirect(url_for('admin_login'))

@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    return render_template('admin.html')

# ---------------------------------------------------------------------------
# ADMIN API ROUTES
# ---------------------------------------------------------------------------

@app.route('/api/admin/students')
@admin_required
def api_admin_students():
    """Return all students without password_hash."""
    students = _read_json('students.json')
    safe = []
    for s in students.values():
        safe.append({
            'id': s.get('id'),
            'name': s.get('name'),
            'email': s.get('email'),
            'disability_type': s.get('disability_type'),
            'dob': s.get('dob'),
            'phone': s.get('phone'),
            'registered_at': s.get('registered_at'),
        })
    # Sort by registration date descending
    safe.sort(key=lambda x: x.get('registered_at', ''), reverse=True)
    return jsonify(safe)

@app.route('/api/admin/questions')
@admin_required
def api_admin_questions():
    """Return current questions (with correct answers, admin view)."""
    return jsonify(_read_json('questions.json'))

@app.route('/api/admin/edit_student', methods=['POST'])
@admin_required
def api_admin_edit_student():
    data = request.json
    student_id = data.get('id')
    if not student_id:
        return jsonify({'status': 'error', 'message': 'Student ID is required'}), 400
        
    students = _read_json('students.json')
    if student_id not in students:
        return jsonify({'status': 'error', 'message': 'Student not found'}), 404
        
    student = students[student_id]
    if 'name' in data: student['name'] = data['name'].strip()
    if 'email' in data: student['email'] = data['email'].strip().lower()
    if 'disability_type' in data: student['disability_type'] = data['disability_type'].strip()
    if 'dob' in data: student['dob'] = data['dob'].strip()
    if 'phone' in data: student['phone'] = data['phone'].strip()
    
    _write_json('students.json', students)
    return jsonify({'status': 'success', 'message': 'Student updated successfully'})

@app.route('/api/admin/results')
@admin_required
def api_admin_results():
    student_id = request.args.get('student_id')
    results = _read_json('results.json')
    if student_id:
        results = [r for r in results if r.get('student_id') == student_id]
        
    results.sort(key=lambda x: x.get('submitted_at', ''), reverse=True)
    return jsonify(results)

@app.route('/api/admin/upload_questions', methods=['POST'])
@admin_required
def api_admin_upload_questions():
    """
    Accept a PDF or XLSX file and update questions.json.

    XLSX expected columns (case-insensitive):
        question | option_a | option_b | option_c | option_d | correct

    PDF expected format per question block:
        1. Question text
        A) Option A
        B) Option B
        C) Option C
        D) Option D
        Answer: C
    """
    if 'file' not in request.files:
        return jsonify({'status': 'error', 'message': 'No file provided'}), 400

    f = request.files['file']
    filename = f.filename.lower()

    questions = []

    # ---- EXCEL ----
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(f.read()))
            ws = wb.active
            headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            required_cols = {'question', 'option_a', 'option_b', 'option_c', 'option_d', 'correct'}
            missing = required_cols - set(headers)
            if missing:
                return jsonify({'status': 'error', 'message': f'Missing columns: {", ".join(missing)}'}), 400

            col = {h: i for i, h in enumerate(headers)}
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                def cell(name):
                    v = row[col[name]]
                    return str(v).strip() if v is not None else ''
                correct = cell('correct').upper()
                if correct not in ('A', 'B', 'C', 'D'):
                    return jsonify({'status': 'error', 'message': f'Row {row_idx}: "correct" must be A/B/C/D, got "{correct}"'}), 400
                questions.append({
                    'id': row_idx - 1,
                    'question': cell('question'),
                    'options': {
                        'A': cell('option_a'),
                        'B': cell('option_b'),
                        'C': cell('option_c'),
                        'D': cell('option_d'),
                    },
                    'correct': correct
                })
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'Excel parse error: {str(e)}'}), 400

    # ---- PDF ----
    elif filename.endswith('.pdf'):
        try:
            import pdfplumber
            file_bytes = f.read()
            
            # Debug: Save the uploaded PDF to disk
            with open(os.path.join(DATA_DIR, 'latest_upload.pdf'), 'wb') as df:
                df.write(file_bytes)
                
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                text = '\n'.join(page.extract_text() or '' for page in pdf.pages)
                
            # Debug: Save extracted text to disk
            with open(os.path.join(DATA_DIR, 'latest_text.txt'), 'w', encoding='utf-8') as dt:
                dt.write(text)

            # Split into blocks by numbered questions (e.g. "1.", "Q1:", "1)", "Q1.")
            blocks = re.split(r'(?:^|\n)\s*(?=[Qq]?\d+[\.\:\)]\s)', text)
            for block in blocks:
                block = block.strip()
                if not block:
                    continue
                
                # First line: "Q1: Question text" (up to the first option or Answer)
                q_text_match = re.match(r'^[Qq]?\d+[\.\:\)]\s*(.*?)(?=\s+[A-D][\.\:\)]|\s+[Aa]nswer|\Z)', block, re.DOTALL)
                if not q_text_match:
                    continue
                q_text = re.sub(r'\s+', ' ', q_text_match.group(1).strip())

                opts = {}
                # Extract options A, B, C, D
                for opt_char in ['A', 'B', 'C', 'D']:
                    next_opts = ''.join([c for c in ['A','B','C','D'] if c > opt_char])
                    if next_opts:
                        lookahead = f"(?:\\s+[{next_opts}][\\.\\:\\)]|\\s+[Aa]nswer|\\Z)"
                    else:
                        lookahead = r"(?:\s+[Aa]nswer|\Z)"
                    
                    pattern = f"{opt_char}[\\.\\:\\)]\\s*(.*?){lookahead}"
                    m = re.search(pattern, block, re.DOTALL | re.IGNORECASE)
                    if m:
                        opts[opt_char] = re.sub(r'\s+', ' ', m.group(1).strip())
                    else:
                        opts[opt_char] = ''
                
                # Extract Answer
                answer_match = re.search(r'[Aa]nswer[^A-D]*(?P<ans>[A-D])', block)
                answer = answer_match.group('ans').upper() if answer_match else ''

                if len([v for v in opts.values() if v]) < 2 or not answer:
                    continue  # skip malformed blocks (at least 2 options and an answer)

                questions.append({
                    'id': len(questions) + 1,
                    'question': q_text,
                    'options': opts,
                    'correct': answer
                })
        except Exception as e:
            return jsonify({'status': 'error', 'message': f'PDF parse error: {str(e)}'}), 400

    else:
        return jsonify({'status': 'error', 'message': 'Unsupported file type. Upload .xlsx or .pdf'}), 400

    if not questions:
        return jsonify({'status': 'error', 'message': 'No valid questions found in the file'}), 400

    _write_json('questions.json', questions)
    return jsonify({'status': 'success', 'message': f'{len(questions)} questions uploaded successfully', 'count': len(questions)})

# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    socketio.run(app, debug=True, allow_unsafe_werkzeug=True, port=5000)